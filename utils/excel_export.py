"""Build an in-memory Excel model summary for a screened deal."""

from datetime import date
from io import BytesIO


def export_excel_model(
    extracted: dict,
    underwriter: dict,
    risk: dict,
    property_name: str = "Deal",
) -> bytes:
    """
    Build a formatted .xlsx workbook and return the raw bytes.

    Single sheet 'Summary' with five sections:
      1. Property Overview
      2. Underwriting Assumptions
      3. Deal Score
      4. Underwriting Metrics
      5. Triggered Risk Flags
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise RuntimeError(f"openpyxl not installed: {exc}")

    def _get(field: str):
        entry = extracted.get(field)
        return entry.get("value") if isinstance(entry, dict) else None

    def _dollar(v):
        return f"${v:,.0f}" if v is not None else "—"

    def _pct(v, decimals=2):
        return f"{v:.{decimals}f}%" if v is not None else "—"

    def _x(v):
        return f"{v:.2f}x" if v is not None else "—"

    # ── data ──────────────────────────────────────────────────────────────────
    loan = underwriter.get("loan_assumptions", {})
    metrics = underwriter.get("metrics", {})
    ds = underwriter.get("deal_score", {})
    flags = [f for f in risk.get("flags", []) if f.get("triggered")]

    prop_rows = [
        ("Property Name",    _get("property_name") or "—"),
        ("Location",         _get("location") or "—"),
        ("Asset Class",      _get("asset_class") or "—"),
        ("Units / SF",       str(_get("units")) if _get("units") else "—"),
        ("Asking Price",     _dollar(_get("asking_price"))),
        ("Year Built",       str(int(_get("year_built"))) if _get("year_built") else "—"),
        ("Occupancy %",      _pct(_get("occupancy_pct"), 1)),
        ("NOI (T-12)",       _dollar(_get("noi_t12"))),
        ("Asking Cap Rate",  _pct(_get("asking_cap_rate"))),
    ]

    assump_rows = [
        ("LTV",                 f"{loan.get('ltv', 0)*100:.0f}%"),
        ("Interest Rate",       f"{loan.get('rate', 0)*100:.1f}%"),
        ("Amortization",        f"{loan.get('amortization_years', '—')} years"),
        ("Hold Period",         f"{loan.get('hold_period_years', '—')} years"),
        ("Loan Amount",         _dollar(loan.get("loan_amount"))),
        ("Equity Invested",     _dollar(loan.get("equity_invested"))),
        ("Annual Debt Service", _dollar(loan.get("annual_debt_service"))),
    ]

    score_rows = [(
        f"{ds.get('score', 0):.0f} / 100",
        ds.get("label", "—"),
        ds.get("explanation", "—"),
    )]

    METRIC_ORDER = [
        ("cap_rate_inplace",  "Cap Rate (T-12)"),
        ("cap_rate_proforma", "Cap Rate (Proforma)"),
        ("dscr",              "DSCR"),
        ("debt_yield",        "Debt Yield"),
        ("cash_on_cash",      "Cash-on-Cash"),
    ]
    metrics_rows = [
        (label, metrics.get(k, {}).get("formatted", "—"),
         metrics.get(k, {}).get("benchmark", "—"),
         metrics.get(k, {}).get("status", "—").upper())
        for k, label in METRIC_ORDER
    ]

    flag_rows = [
        (f.get("flag_name", "—"), f.get("severity", "—").upper(), f.get("detail", "—"))
        for f in flags
    ] or [("No risk flags triggered", "", "")]

    # ── build workbook ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 55

    HEADER_FILL = PatternFill(fgColor="16213e", fill_type="solid")
    SUBHDR_FILL = PatternFill(fgColor="d9d9d9", fill_type="solid")
    WHITE_FONT  = Font(bold=True, color="FFFFFF")
    BOLD_FONT   = Font(bold=True)

    def _section(title: str, col_headers: list, rows: list) -> None:
        # Section title
        ws.append([title])
        r = ws.max_row
        ws.merge_cells(f"A{r}:D{r}")
        c = ws.cell(r, 1)
        c.fill = HEADER_FILL
        c.font = WHITE_FONT
        c.alignment = Alignment(vertical="center")

        # Column headers
        ws.append(col_headers)
        r = ws.max_row
        for ci in range(1, len(col_headers) + 1):
            cell = ws.cell(r, ci)
            cell.fill = SUBHDR_FILL
            cell.font = BOLD_FONT

        # Data rows
        for row in rows:
            ws.append(list(row))

        ws.append([])  # blank separator

    # Title block
    ws.append([f"PD Deal Screening Model — {property_name}"])
    ws.cell(ws.max_row, 1).font = Font(bold=True, size=14)
    ws.append([f"Generated: {date.today().isoformat()}"])
    ws.append([])

    _section("PROPERTY OVERVIEW",      ["Field", "Value"],                       prop_rows)
    _section("UNDERWRITING ASSUMPTIONS", ["Parameter", "Value"],                  assump_rows)
    _section("DEAL SCORE",             ["Score", "Label", "Explanation"],         score_rows)
    _section("UNDERWRITING METRICS",   ["Metric", "Value", "Benchmark", "Status"], metrics_rows)
    _section("TRIGGERED RISK FLAGS",   ["Flag", "Severity", "Detail"],            flag_rows)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
